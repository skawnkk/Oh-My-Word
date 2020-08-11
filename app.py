from pymongo import MongoClient
from bs4 import BeautifulSoup
import requests
from openpyxl import load_workbook
from flask import Flask, render_template, jsonify, request
app = Flask(__name__, template_folder='templates')

# 크롤링

# 파이몽고
client = MongoClient('localhost', 27017)
# client = MongoClient('mongodb://test:test@localhost', 27017)
db = client.dbsparta




# <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<INDEX HTML을 주는 부분
@app.route('/')
def home():
    return render_template('main.html')

# API 역할을 하는 부분

@app.route('/plus', methods=['POST'])
def plus():
    word_receive = request.form['word_give']
    title_receive = request.form['title_give']
    plus_receive =request.form['plus_give']
    
    db.wordcards.update_one({'word':word_receive,'title':title_receive},{'$set': {'plus':plus_receive}})

    return jsonify({'result': 'success', 'msg': '등록완료!'})



@app.route('/word', methods=['POST'])
def make_word():
    word_receive = request.form['word_give']
    title_receive = request.form['title_give']
    plus_receive = request.form['plus_give']
    
    url = "http://endic.naver.com/search.nhn?query=" + word_receive
    data = requests.get(url)
    soup = BeautifulSoup(data.text, 'html.parser')

    # print(data.text)

    result=""

    try:
        result += soup.find('dl', {'class': 'list_e2'}).find(
            'dd').find('span', {'class': 'fnt_k05'}).get_text()

    except:
        result = "네이버 사전에 등재되어 있지 않습니다."

    doc = {
        'title': title_receive,
        'word': word_receive,
        'meaning': result,
        'plus':plus_receive
        
    }

    db.wordcards.insert_one(doc)

    return jsonify({'result': 'success', 'msg': '등록완료!'})


@app.route('/card', methods=['GET'])
def listing():
    title = request.args.get('title')
    wordcards = list(db.wordcards.find({'title': title}, {'_id': 0}))
    return jsonify({'result': 'success', 'wordcards': wordcards})


@app.route('/nocard', methods=['post'])
def delete():
    word_receive = request.form['word_give']
    title_receive = request.form['title_give']

    db.wordcards.delete_one({'word': word_receive, 'title': title_receive})

    return jsonify({'result': 'success', 'msg': '삭제되었습니다.'})


@app.route('/savetitle', methods=['post'])
def savetitle():
    title_receive = request.form['title_give']
    db.wordcards.update_many({'title': ''}, {'$set': {'title': title_receive}})
    return jsonify({'result': 'success'})


@app.route('/excel', methods=['POST'])
def excel():
    print('excel')
    title_receive = request.form['title_give']
    cards = list(db.wordcards.find({'title': title_receive}, {'_id': 0}))
    # print(cards)
  
    work_book = load_workbook('test.xlsx')
    work_sheet = work_book['prac']

    row = 2
    for card in cards:
        work_sheet.cell(row=row, column=2, value=card['title'])
        work_sheet.cell(row=row, column=3, value=card['word'])
        work_sheet.cell(row=row, column=4, value=card['meaning'])
        row += 1
        
        work_book.save('test.xlsx')
        
    return jsonify({'result': 'success', 'msg': '엑셀파일을 확인해보세요.'})


# <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<MAIN.HTML화면 연결하기
@app.route('/listpage')
def makenew():
    return render_template('index.html')

@app.route('/listpage2')
def wordlist():
    pagetitle = request.args.get('pagetitle')
    pagetitle = pagetitle.replace("'", "")
    return render_template('index2.html', pagetitle=pagetitle)


@app.route('/folderlist', methods=['GET'])
def folderlist():
    folderlists = db.wordcards.distinct("title")
    return jsonify({'result': 'success', 'folderlists': folderlists})


@app.route('/deletefolder', methods=['POST'])
def deletefolder():
  
    foldertitle_receive = request.form('foldertitle_give') 
    db.wordcards.deleteMany({'title':foldertitle_receive},{'_id':0})
    return jsonify({'result': 'success', 'msg': '삭제되었습니다.'})


if __name__ == '__main__':
    app.run('0.0.0.0', port=5000, debug=True)
